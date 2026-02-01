"""
Excel export for financial model.
Builds the complete financial model workbook with variable costs pre-filled.
"""

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class FinancialData:
    """Data needed to build the financial model export."""

    period: datetime
    call_count: int
    total_minutes: float
    llm_cost: float
    stt_cost: float
    tts_cost: float
    telephony_cost: float
    customer_data: List[Tuple[str, float, int]]  # (name, cost, call_count)


class FinancialModelExporter:
    """Builds Excel financial model workbook."""

    # Styles
    HEADER_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=11)
    SECTION_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    CURRENCY_FMT = '"$"#,##0.00'
    PCT_FMT = "0.0%"
    THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

    def __init__(self, data: FinancialData):
        self.data = data
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active
        self.ws.title = "Financial Model"
        self.row = 1

        # Track row numbers for formulas
        self.variable_total_row = 0
        self.infra_total_row = 0
        self.revenue_total_row = 0
        self.income_revenue_row = 0
        self.gross_profit_row = 0
        self.net_income_row = 0

    def build(self) -> BytesIO:
        """Build the complete workbook and return as BytesIO."""
        self._add_title()
        self._add_variable_costs()
        self._add_unit_economics()
        self._add_infrastructure_costs()
        self._add_revenue()
        self._add_income_statement()
        self._add_runway()
        self._add_key_metrics()
        self._set_column_widths()

        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _cell(self, col: int, value, font=None, fill=None, number_format=None, border=None):
        """Helper to set cell value and styles."""
        cell = self.ws.cell(row=self.row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        if border:
            cell.border = border
        return cell

    def _add_section_header(self, title: str):
        """Add a section header row."""
        self._cell(1, title, font=self.HEADER_FONT, fill=self.SECTION_FILL)
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
        self.row += 1

    def _add_column_headers(self, headers: List[str]):
        """Add column header row."""
        for col, header in enumerate(headers, 1):
            self._cell(col, header, font=self.SECTION_FONT, border=self.THIN_BORDER)
        self.row += 1

    def _add_title(self):
        """Add title and metadata."""
        self._cell(1, "OptimalBot Financial Model", font=Font(bold=True, size=16))
        self.row += 1
        self._cell(1, f"Period: {self.data.period.strftime('%B %Y')}")
        self.row += 1
        self._cell(1, f"Generated: {self.data.period.strftime('%Y-%m-%d %H:%M UTC')}")
        self.row += 2

    def _add_variable_costs(self):
        """Add variable costs section."""
        self._add_section_header("VARIABLE COSTS (Per-Call) - Auto-filled")
        self._add_column_headers(["Component", "MTD Cost", "% of Total"])

        variable_start = self.row
        items = [
            ("LLM (OpenAI/Groq)", round(self.data.llm_cost, 4)),
            ("STT (Deepgram)", round(self.data.stt_cost, 4)),
            ("TTS (Cartesia)", round(self.data.tts_cost, 4)),
            ("Telephony (Daily)", round(self.data.telephony_cost, 4)),
        ]
        for name, cost in items:
            self._cell(1, name)
            self._cell(2, cost, number_format=self.CURRENCY_FMT)
            self.row += 1
        variable_end = self.row - 1

        # Total row
        self._cell(1, "Total Variable Costs", font=self.SECTION_FONT)
        self._cell(2, f"=SUM(B{variable_start}:B{variable_end})",
                   font=self.SECTION_FONT, number_format=self.CURRENCY_FMT)
        self.variable_total_row = self.row

        # Add percentage formulas
        for r in range(variable_start, variable_end + 1):
            self.ws.cell(row=r, column=3,
                         value=f"=IF(B{self.variable_total_row}=0,0,B{r}/B{self.variable_total_row})"
                         ).number_format = self.PCT_FMT
        self.row += 2

    def _add_unit_economics(self):
        """Add unit economics section."""
        self._cell(1, "Unit Economics", font=self.SECTION_FONT)
        self.row += 1

        self._cell(1, "Total Calls (MTD)")
        self._cell(2, self.data.call_count)
        calls_row = self.row
        self.row += 1

        self._cell(1, "Total Minutes (MTD)")
        self._cell(2, round(self.data.total_minutes, 2))
        minutes_row = self.row
        self.row += 1

        self._cell(1, "Avg Cost/Call")
        self._cell(2, f"=IF(B{calls_row}=0,0,B{self.variable_total_row}/B{calls_row})",
                   number_format=self.CURRENCY_FMT)
        self.row += 1

        self._cell(1, "Avg Cost/Minute")
        self._cell(2, f"=IF(B{minutes_row}=0,0,B{self.variable_total_row}/B{minutes_row})",
                   number_format=self.CURRENCY_FMT)
        self.row += 2

    def _add_infrastructure_costs(self):
        """Add infrastructure costs section."""
        self._add_section_header("INFRASTRUCTURE COSTS (Monthly) - Enter manually")
        self._add_column_headers(["Service", "Monthly Cost", "Notes"])

        infra_start = self.row
        items = [
            ("Fly.io", "Backend hosting"),
            ("Vercel", "Frontend hosting"),
            ("Pipecat Cloud", "Voice pipeline"),
            ("MongoDB Atlas", "Database (HIPAA)"),
            ("Daily.co", "Telephony platform"),
            ("Langfuse", "Observability"),
            ("Other", ""),
        ]
        for name, note in items:
            self._cell(1, name)
            self._cell(2, 0, number_format=self.CURRENCY_FMT)
            self._cell(3, note)
            self.row += 1
        infra_end = self.row - 1

        self._cell(1, "Total Infrastructure", font=self.SECTION_FONT)
        self._cell(2, f"=SUM(B{infra_start}:B{infra_end})",
                   font=self.SECTION_FONT, number_format=self.CURRENCY_FMT)
        self.infra_total_row = self.row
        self.row += 2

    def _add_revenue(self):
        """Add revenue section."""
        self._add_section_header("REVENUE - Enter manually")
        self._add_column_headers(["Customer", "MRR", "COGS (auto)"])

        revenue_start = self.row
        customers = (
            self.data.customer_data[:5]
            if self.data.customer_data
            else [("Customer 1", 0, 0), ("Customer 2", 0, 0), ("Customer 3", 0, 0)]
        )
        for name, cogs, _ in customers:
            self._cell(1, name)
            self._cell(2, 0, number_format=self.CURRENCY_FMT)  # MRR - manual
            self._cell(3, round(cogs, 4), number_format=self.CURRENCY_FMT)  # COGS - from data
            self.row += 1

        # Add empty rows for more customers
        for _ in range(max(0, 5 - len(customers))):
            self._cell(2, 0, number_format=self.CURRENCY_FMT)
            self._cell(3, 0, number_format=self.CURRENCY_FMT)
            self.row += 1
        revenue_end = self.row - 1

        self._cell(1, "Total MRR", font=self.SECTION_FONT)
        self._cell(2, f"=SUM(B{revenue_start}:B{revenue_end})",
                   font=self.SECTION_FONT, number_format=self.CURRENCY_FMT)
        self.revenue_total_row = self.row
        self.row += 1

        self._cell(1, "ARR")
        self._cell(2, f"=B{self.revenue_total_row}*12", number_format=self.CURRENCY_FMT)
        self.row += 2

    def _add_income_statement(self):
        """Add income statement section."""
        self._add_section_header("INCOME STATEMENT (Monthly)")
        self._add_column_headers(["Line Item", "Amount", "% of Revenue"])

        # Revenue
        self._cell(1, "Revenue")
        self._cell(2, f"=B{self.revenue_total_row}", number_format=self.CURRENCY_FMT)
        self.income_revenue_row = self.row
        self.row += 1

        # COGS
        self._cell(1, "COGS (Variable Costs)")
        self._cell(2, f"=B{self.variable_total_row}", number_format=self.CURRENCY_FMT)
        self._cell(3, f"=IF(B{self.income_revenue_row}=0,0,B{self.row}/B{self.income_revenue_row})",
                   number_format=self.PCT_FMT)
        cogs_row = self.row
        self.row += 1

        # Gross Profit
        self._cell(1, "Gross Profit", font=self.SECTION_FONT)
        self._cell(2, f"=B{self.income_revenue_row}-B{cogs_row}",
                   font=self.SECTION_FONT, number_format=self.CURRENCY_FMT)
        self._cell(3, f"=IF(B{self.income_revenue_row}=0,0,B{self.row}/B{self.income_revenue_row})",
                   number_format=self.PCT_FMT)
        self.gross_profit_row = self.row
        self.row += 1

        # Operating Expenses
        self._cell(1, "Operating Expenses (Infra)")
        self._cell(2, f"=B{self.infra_total_row}", number_format=self.CURRENCY_FMT)
        self._cell(3, f"=IF(B{self.income_revenue_row}=0,0,B{self.row}/B{self.income_revenue_row})",
                   number_format=self.PCT_FMT)
        opex_row = self.row
        self.row += 1

        # Net Income
        self._cell(1, "Net Income", font=Font(bold=True, size=12))
        self._cell(2, f"=B{self.gross_profit_row}-B{opex_row}",
                   font=Font(bold=True, size=12), number_format=self.CURRENCY_FMT)
        self._cell(3, f"=IF(B{self.income_revenue_row}=0,0,B{self.row}/B{self.income_revenue_row})",
                   number_format=self.PCT_FMT)
        self.net_income_row = self.row
        self.row += 2

    def _add_runway(self):
        """Add runway section."""
        self._add_section_header("RUNWAY")

        self._cell(1, "Cash on Hand")
        self._cell(2, 0, number_format=self.CURRENCY_FMT)
        cash_row = self.row
        self.row += 1

        self._cell(1, "Monthly Burn")
        self._cell(2, f"=IF(B{self.net_income_row}<0,-B{self.net_income_row},0)",
                   number_format=self.CURRENCY_FMT)
        burn_row = self.row
        self.row += 1

        self._cell(1, "Runway (Months)", font=self.SECTION_FONT)
        self._cell(2, f'=IF(B{burn_row}=0,"Profitable",B{cash_row}/B{burn_row})',
                   font=self.SECTION_FONT)
        self.row += 2

    def _add_key_metrics(self):
        """Add key metrics section."""
        self._add_section_header("KEY METRICS")

        metrics = [
            ("Gross Margin", f"=IF(B{self.income_revenue_row}=0,0,B{self.gross_profit_row}/B{self.income_revenue_row})"),
            ("Net Margin", f"=IF(B{self.income_revenue_row}=0,0,B{self.net_income_row}/B{self.income_revenue_row})"),
            ("Variable Cost % of Revenue", f"=IF(B{self.income_revenue_row}=0,0,B{self.variable_total_row}/B{self.income_revenue_row})"),
            ("Infra Cost % of Revenue", f"=IF(B{self.income_revenue_row}=0,0,B{self.infra_total_row}/B{self.income_revenue_row})"),
        ]
        for name, formula in metrics:
            self._cell(1, name)
            self._cell(2, formula, number_format=self.PCT_FMT)
            self.row += 1

    def _set_column_widths(self):
        """Set column widths."""
        self.ws.column_dimensions["A"].width = 35
        self.ws.column_dimensions["B"].width = 18
        self.ws.column_dimensions["C"].width = 20
